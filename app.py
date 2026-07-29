import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

st.set_page_config(
    page_title="POI Hit Report Processor | Dev by Ashaan",
    page_icon="📍",
    layout="wide"
)

# ------------------------------------------------------------------------------
# AESTHETIC UI STYLING
# ------------------------------------------------------------------------------
st.markdown("""
    <style>
    .header-card {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 10px 25px rgba(0,0,0,0.2);
        margin-bottom: 25px;
        color: white;
        border: 1px solid #334155;
    }
    .header-title {
        font-family: 'Segoe UI', sans-serif;
        font-size: 26px;
        font-weight: 700;
        text-align: center;
        letter-spacing: 0.5px;
        color: #F8FAFC;
    }
    .header-sub {
        font-family: 'Segoe UI', sans-serif;
        text-align: center;
        color: #94A3B8;
        font-size: 14px;
        margin-top: 4px;
    }
    .header-sub b {
        color: #38BDF8;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="header-card">
    <div class="header-title">📍 POI HIT & CONTAINER AUDIT DASHBOARD</div>
    <div class="header-sub">Developed by: <b>Muhammad Ashaan</b> | Bellanix Tech</div>
</div>
""", unsafe_allow_html=True)

# ------------------------------------------------------------------------------
# CITY SELECTION & FILE UPLOADER
# ------------------------------------------------------------------------------
st.markdown("### 📍 City & File Selection")

c_col1, c_col2 = st.columns([1, 2])

with c_col1:
    city_option = st.selectbox(
        "Select City Name:",
        options=["Kamoke", "Nowshera Virkan", "Gujranwala", "Other (Custom)"]
    )
    if city_option == "Other (Custom)":
        selected_city = st.text_input("Enter City Name:", value="Kamoke").strip().upper()
    else:
        selected_city = city_option.upper()

with c_col2:
    uploaded_file = st.file_uploader("📥 Upload POI Hit Report (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
        
        # Detect Metadata
        generated_at = str(df_raw.iloc[1, 1]).strip() if pd.notna(df_raw.iloc[1, 1]) else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_period = str(df_raw.iloc[2, 1]).strip() if pd.notna(df_raw.iloc[2, 1]) else "N/A"

        # Find Header Row
        header_row_idx = -1
        for r in range(min(10, len(df_raw))):
            row_str = [str(x).strip().upper() for x in df_raw.iloc[r].tolist()]
            if 'SER#' in row_str and 'CONT#' in row_str and 'VEHICLE' in row_str:
                header_row_idx = r
                break

        if header_row_idx == -1:
            st.error("❌ Invalid POI Hit Report format. 'Ser#', 'Cont#', and 'Vehicle' columns were not found.")
        else:
            headers = [str(x).strip() for x in df_raw.iloc[header_row_idx].tolist()]
            df_data = df_raw.iloc[header_row_idx + 1:].copy()
            df_data.columns = headers

            # Clean Status
            df_data['Vehicle_Clean'] = df_data['Vehicle'].astype(str).str.strip()
            
            not_picked_df = df_data[df_data['Vehicle_Clean'].str.upper() == 'NOT PICKED!'].copy()
            picked_df = df_data[df_data['Vehicle_Clean'].str.upper() != 'NOT PICKED!'].copy()

            total_entries = len(df_data)
            not_picked_count = len(not_picked_df)
            picked_count = len(picked_df)
            pick_percentage = (picked_count / total_entries * 100) if total_entries > 0 else 0

            # Exclude empty columns (#, Time, Proximity, PITB StopPointId) for Not Picked list
            empty_cols_to_remove = ['#', 'Time', 'Proximity', 'PITB StopPointId', 'Vehicle_Clean']
            clean_not_picked_cols = [c for c in headers if c not in empty_cols_to_remove]
            
            # Keep ALL 50 rows (No deduplication) & Assign Sequential Ser# (1, 2, 3, 4...)
            display_not_picked = not_picked_df[clean_not_picked_cols].copy().reset_index(drop=True)
            display_not_picked['Ser#'] = range(1, len(display_not_picked) + 1)

            display_picked = picked_df[headers].copy().reset_index(drop=True)
            display_picked['Ser#'] = range(1, len(display_picked) + 1)

            # ------------------------------------------------------------------
            # WEB PORTAL OVERVIEW
            # ------------------------------------------------------------------
            st.markdown("---")
            st.markdown("### 📊 Web Portal Overview & Status Summary")
            
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Selected City", selected_city)
            m2.metric("Total POI Locations", total_entries)
            m3.metric("Picked Status", f"{picked_count} Picked", delta=f"{pick_percentage:.1f}% Completed")
            m4.metric("Not Picked / Pending", f"{not_picked_count} Pending", delta=f"-{100-pick_percentage:.1f}% Remaining", delta_color="inverse")

            st.info(f"🕒 **Generated Date:** {generated_at}  |  📅 **Time Period:** {time_period}")

            # Status Breakdown Table
            st.markdown("#### 📋 Container Picking Status Summary")
            status_summary_df = pd.DataFrame([
                {"Status": "🟢 Picked Containers", "Count": picked_count, "Percentage": f"{pick_percentage:.1f}%"},
                {"Status": "🔴 Not Picked / Pending", "Count": not_picked_count, "Percentage": f"{100-pick_percentage:.1f}%"},
                {"Status": "🔹 Total Locations", "Count": total_entries, "Percentage": "100.0%"}
            ])
            st.table(status_summary_df)

            # Preview Table
            st.markdown("---")
            st.markdown(f"### ⚠️ NOT PICKED CONTAINERS OR SKIPS REPORT ({not_picked_count} LOCATIONS)")
            
            st.dataframe(
                display_not_picked[['Ser#', 'Cont#', 'Vehicle', 'Latitude', 'Longitude']],
                use_container_width=True,
                height=340
            )

            # ------------------------------------------------------------------
            # EXCEL GENERATOR
            # ------------------------------------------------------------------
            wb = openpyxl.Workbook()
            
            font_title = Font(name='Segoe UI', size=15, bold=True, color='FFFFFF')
            font_sub = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            font_meta = Font(name='Segoe UI', size=10, italic=True, color='E2E8F0')
            font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            font_body = Font(name='Segoe UI', size=10, color='1E293B')
            font_alert = Font(name='Segoe UI', size=10, bold=True, color='991B1B')
            
            fill_title = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
            fill_sub = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid')
            fill_meta = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
            fill_header = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
            fill_alt = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            fill_not_picked = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin', color='CBD5E0'),
                right=Side(style='thin', color='CBD5E0'),
                top=Side(style='thin', color='CBD5E0'),
                bottom=Side(style='thin', color='CBD5E0')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # --------------------------------------------------
            # SHEET 1: NOT PICKED CONTAINERS
            # --------------------------------------------------
            ws_not_picked = wb.active
            ws_not_picked.title = "Not Picked Containers"
            num_np_cols = len(clean_not_picked_cols)

            # Row 1: <CITY> POI HIT REPORT
            ws_not_picked.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_np_cols)
            cell1 = ws_not_picked.cell(row=1, column=1, value=f"{selected_city} POI HIT REPORT")
            cell1.font = font_title
            cell1.fill = fill_title
            cell1.alignment = center_align
            ws_not_picked.row_dimensions[1].height = 36

            # Row 2: NOT PICKED CONTAINERS OR SKIPS REPORT (<EXACT_COUNT> LOCATIONS)
            ws_not_picked.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_np_cols)
            cell2 = ws_not_picked.cell(row=2, column=1, value=f"NOT PICKED CONTAINERS OR SKIPS REPORT ({not_picked_count} LOCATIONS)")
            cell2.font = font_sub
            cell2.fill = fill_sub
            cell2.alignment = center_align
            ws_not_picked.row_dimensions[2].height = 28

            # Row 3: Generated Date, Period & Right Corner Developed by
            ws_not_picked.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_np_cols)
            cell3 = ws_not_picked.cell(row=3, column=1, value=f"Generated: {generated_at}   |   Period: {time_period}   |   Developed by: Muhammad Ashaan")
            cell3.font = font_meta
            cell3.fill = fill_meta
            cell3.alignment = right_align
            ws_not_picked.row_dimensions[3].height = 22

            # Row 4: Headers
            ws_not_picked.row_dimensions[4].height = 28
            for col_i, h_text in enumerate(clean_not_picked_cols, 1):
                c = ws_not_picked.cell(row=4, column=col_i, value=h_text)
                c.font = font_header
                c.fill = fill_header
                c.alignment = center_align
                c.border = thin_border

            # Data Rows
            row_idx = 5
            for _, r in display_not_picked[clean_not_picked_cols].iterrows():
                ws_not_picked.row_dimensions[row_idx].height = 20
                for col_i, val in enumerate(r, 1):
                    c = ws_not_picked.cell(row=row_idx, column=col_i, value=val)
                    c.alignment = left_align if col_i == 2 else center_align
                    c.border = thin_border
                    c.font = font_alert if col_i == 3 else font_body
                    if col_i == 3:
                        c.fill = fill_not_picked
                    elif row_idx % 2 == 0:
                        c.fill = fill_alt
                row_idx += 1

            # Column Widths (Ser# strictly width 9)
            for col in ws_not_picked.columns:
                col_letter = get_column_letter(col[0].column)
                if col_letter == 'A':
                    ws_not_picked.column_dimensions[col_letter].width = 9
                else:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws_not_picked.column_dimensions[col_letter].width = max(max_len + 4, 15)

            # --------------------------------------------------
            # SHEET 2: SUCCESSFUL PICKED HITS
            # --------------------------------------------------
            ws_hits = wb.create_sheet(title="Picked Hits")
            num_cols = len(headers)

            # Row 1
            ws_hits.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            c1 = ws_hits.cell(row=1, column=1, value=f"{selected_city} SUCCESSFUL POI HITS REPORT")
            c1.font = font_title
            c1.fill = fill_title
            c1.alignment = center_align
            ws_hits.row_dimensions[1].height = 36

            # Row 2
            ws_hits.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            c2 = ws_hits.cell(row=2, column=1, value=f"SUCCESSFUL PICKED CONTAINERS ({picked_count} LOCATIONS)")
            c2.font = font_sub
            c2.fill = PatternFill(start_color='15803D', end_color='15803D', fill_type='solid')
            c2.alignment = center_align
            ws_hits.row_dimensions[2].height = 28

            # Row 3
            ws_hits.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
            c3 = ws_hits.cell(row=3, column=1, value=f"Generated: {generated_at}   |   Period: {time_period}   |   Developed by: Muhammad Ashaan")
            c3.font = font_meta
            c3.fill = fill_meta
            c3.alignment = right_align
            ws_hits.row_dimensions[3].height = 22

            # Row 4
            ws_hits.row_dimensions[4].height = 28
            for col_i, h_text in enumerate(headers, 1):
                c = ws_hits.cell(row=4, column=col_i, value=h_text)
                c.font = font_header
                c.fill = fill_header
                c.alignment = center_align
                c.border = thin_border

            # Data Rows
            row_idx = 5
            for _, r in display_picked[headers].iterrows():
                ws_hits.row_dimensions[row_idx].height = 20
                for col_i, val in enumerate(r, 1):
                    c = ws_hits.cell(row=row_idx, column=col_i, value=val if pd.notna(val) else "")
                    c.alignment = left_align if col_i == 2 else center_align
                    c.border = thin_border
                    c.font = font_body
                    if row_idx % 2 == 0:
                        c.fill = fill_alt
                row_idx += 1

            # Column Widths (Ser# strictly width 9)
            for col in ws_hits.columns:
                col_letter = get_column_letter(col[0].column)
                if col_letter == 'A':
                    ws_hits.column_dimensions[col_letter].width = 9
                else:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws_hits.column_dimensions[col_letter].width = max(max_len + 4, 14)

            # --------------------------------------------------
            # DOWNLOAD
            # --------------------------------------------------
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            st.markdown("---")
            st.download_button(
                label=f"📥 Download {selected_city} Executive POI Report (.xlsx)",
                data=excel_buffer,
                file_name=f"{selected_city}_POI_Hit_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"❌ Error processing report: {str(e)}")
